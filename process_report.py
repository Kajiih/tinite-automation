"""
Amazon VAT Transaction Report - FC_Transfer Price Automation

Production-grade automation to populate unit cost (Column T: COST_PRICE_OF_ITEMS)
and line total (Column U: PRICE_OF_ITEMS_AMT_VAT_EXCL) for FC_TRANSFER rows in
Amazon VAT reports based on an Excel price catalog.

Calculates and exports cross-border transfer sums by Departure Country x Arrival Country.
Supports both single CSV file processing and batch folder processing.
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from collections.abc import Mapping, MutableMapping, MutableSequence, Sequence, Set
from dataclasses import dataclass, field
from enum import StrEnum
from pathlib import Path


class TransactionType(StrEnum):
    """Supported transaction types in Amazon VAT transaction reports."""
    FC_TRANSFER = "FC_TRANSFER"
    SALE = "SALE"
    REFUND = "REFUND"
    RETURN = "RETURN"


class ColumnHeader(StrEnum):
    """Standard column header names in Amazon VAT transaction reports."""
    TRANSACTION_TYPE = "TRANSACTION_TYPE"
    ASIN = "ASIN"
    QUANTITY = "QTY"
    COST_PRICE_OF_ITEMS = "COST_PRICE_OF_ITEMS"
    PRICE_OF_ITEMS_AMT_VAT_EXCL = "PRICE_OF_ITEMS_AMT_VAT_EXCL"
    TOTAL_PRICE_OF_ITEMS_AMT_VAT_EXCL = "TOTAL_PRICE_OF_ITEMS_AMT_VAT_EXCL"
    TOTAL_ACTIVITY_VALUE_AMT_VAT_EXCL = "TOTAL_ACTIVITY_VALUE_AMT_VAT_EXCL"
    DEPARTURE_COUNTRY = "DEPARTURE_COUNTRY"
    ARRIVAL_COUNTRY = "ARRIVAL_COUNTRY"


UNKNOWN_COUNTRY_CODE: str = "UNKNOWN"
DEFAULT_ENCODING: str = "utf-8-sig"
CSV_LINE_TERMINATOR: str = "\r\n"


@dataclass(frozen=True, order=True)
class RouteKey:
    """Represents a cross-border transfer route (Departure Country -> Arrival Country)."""
    departure_country: str
    arrival_country: str

    def __str__(self) -> str:
        return f"{self.departure_country} -> {self.arrival_country}"


@dataclass
class RouteMetric:
    """Accumulator for cross-border transfer statistics along a specific route."""
    transfer_count: int = 0
    total_quantity: float = 0.0
    total_amount_eur: float = 0.0

    def add_transfer(self, quantity: float, amount_eur: float) -> None:
        """Add a single transfer event to the accumulated route metrics."""
        self.transfer_count += 1
        self.total_quantity += quantity
        self.total_amount_eur += amount_eur

    def merge_metrics(self, other: RouteMetric) -> None:
        """Merge metrics from another route accumulator."""
        self.transfer_count += other.transfer_count
        self.total_quantity += other.total_quantity
        self.total_amount_eur += other.total_amount_eur


@dataclass
class FileProcessingResult:
    """Detailed summary statistics for a single processed VAT report."""
    report_path: Path
    output_path: Path
    summary_path: Path | None
    total_rows: int
    fc_transfer_count: int
    fc_transfer_updated: int
    missing_asins: Sequence[str]
    missing_rows_count: int
    total_value_added: float
    route_statistics: Mapping[RouteKey, RouteMetric]


@dataclass
class BatchProcessingResult:
    """Consolidated summary statistics for an entire batch of processed VAT reports."""
    input_directory: Path
    output_directory: Path
    batch_summary_path: Path
    price_catalog_path: Path
    catalog_size: int
    files_count: int
    file_results: Sequence[FileProcessingResult]
    grand_total_rows: int = 0
    grand_fc_transfers: int = 0
    grand_fc_updated: int = 0
    grand_value_added: float = 0.0
    all_missing_asins: Set[str] = field(default_factory=set)
    consolidated_routes: MutableMapping[RouteKey, RouteMetric] = field(
        default_factory=lambda: defaultdict(RouteMetric)
    )


def load_price_catalog(price_catalog_path: Path) -> Mapping[str, float]:
    """
    Load ASIN -> unit price mapping from an Excel catalog (.xlsx).
    
    Dynamically scans worksheets to locate the ASIN and unit price columns,
    robustly parsing numbers with commas, periods, and currency symbols.
    """
    import openpyxl

    if not price_catalog_path.exists():
        raise FileNotFoundError(f"Excel price catalog file not found at: {price_catalog_path}")

    workbook = openpyxl.load_workbook(filename=price_catalog_path, data_only=True, read_only=True)
    target_worksheet = None
    asin_column_index: int = 0
    price_column_index: int = 1

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        first_row = next(worksheet.iter_rows(values_only=True), None)
        if not first_row:
            continue

        found_asin_header = False
        for column_index, cell_value in enumerate(first_row):
            if cell_value is None:
                continue
            normalized_header = str(cell_value).strip().lower()
            if "asin" in normalized_header:
                asin_column_index = column_index
                found_asin_header = True
            elif any(
                keyword in normalized_header
                for keyword in ["price cost", "prix d'achat", "cogs", "cost", "price", "prix"]
            ):
                price_column_index = column_index

        if found_asin_header:
            target_worksheet = worksheet
            break

    if target_worksheet is None:
        target_worksheet = workbook.active
        if target_worksheet is None:
            raise ValueError(f"No readable worksheets found in Excel file: {price_catalog_path}")

    price_catalog: MutableMapping[str, float] = {}
    for row_values in target_worksheet.iter_rows(min_row=2, values_only=True):
        if not row_values or len(row_values) <= asin_column_index:
            continue

        raw_asin = row_values[asin_column_index]
        if raw_asin is None:
            continue

        asin = str(raw_asin).strip()
        if not asin or asin.lower() == "asin":
            continue

        raw_price = row_values[price_column_index] if len(row_values) > price_column_index else None
        if raw_price is None:
            continue

        try:
            if isinstance(raw_price, (int, float)):
                price = float(raw_price)
            else:
                sanitized_price_string = (
                    str(raw_price)
                    .replace(",", ".")
                    .replace("€", "")
                    .replace("\xa0", "")
                    .strip()
                )
                price = float(sanitized_price_string)
            price_catalog[asin] = price
        except (ValueError, TypeError):
            continue

    workbook.close()
    if not price_catalog:
        raise ValueError(f"No valid ASIN price entries were extracted from: {price_catalog_path}")

    return price_catalog


def export_country_summary(
    route_statistics: Mapping[RouteKey, RouteMetric],
    output_csv_path: Path,
) -> None:
    """Export Departure Country x Arrival Country route metrics to a clean CSV summary file."""
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)

    header: Sequence[str] = [
        "DEPARTURE_COUNTRY",
        "ARRIVAL_COUNTRY",
        "TRANSFER_COUNT",
        "TOTAL_QTY",
        "TOTAL_AMOUNT_EUR",
    ]

    total_accumulator = RouteMetric()
    output_rows: MutableSequence[Sequence[str]] = [header]

    for route_key in sorted(route_statistics.keys()):
        metric = route_statistics[route_key]
        total_accumulator.merge_metrics(metric)

        quantity_string = (
            f"{int(metric.total_quantity)}"
            if metric.total_quantity.is_integer()
            else f"{metric.total_quantity:.2f}"
        )
        output_rows.append([
            route_key.departure_country,
            route_key.arrival_country,
            str(metric.transfer_count),
            quantity_string,
            f"{metric.total_amount_eur:.2f}",
        ])

    grand_quantity_string = (
        f"{int(total_accumulator.total_quantity)}"
        if total_accumulator.total_quantity.is_integer()
        else f"{total_accumulator.total_quantity:.2f}"
    )
    output_rows.append([
        "TOTAL",
        "",
        str(total_accumulator.transfer_count),
        grand_quantity_string,
        f"{total_accumulator.total_amount_eur:.2f}",
    ])

    with open(output_csv_path, mode="w", encoding=DEFAULT_ENCODING, newline="") as file_handle:
        writer = csv.writer(file_handle, quoting=csv.QUOTE_ALL, lineterminator=CSV_LINE_TERMINATOR)
        writer.writerows(output_rows)


def format_route_table(route_statistics: Mapping[RouteKey, RouteMetric]) -> str:
    """Format route metrics into an aligned, readable terminal table."""
    if not route_statistics:
        return "  No FC_TRANSFER routes recorded.\n"

    table_lines: MutableSequence[str] = [
        f"  {'Route (Depart -> Arrive)':<26} | {'Transfers':<10} | {'Units Moved':<12} | {'Total Value (€)':<15}",
        "  " + "-" * 70,
    ]

    total_accumulator = RouteMetric()
    for route_key in sorted(route_statistics.keys()):
        metric = route_statistics[route_key]
        total_accumulator.merge_metrics(metric)

        quantity_string = (
            f"{int(metric.total_quantity):,d}"
            if metric.total_quantity.is_integer()
            else f"{metric.total_quantity:,.2f}"
        )
        table_lines.append(
            f"  {str(route_key):<26} | {metric.transfer_count:<10,d} | {quantity_string:<12} | €{metric.total_amount_eur:<14,.2f}"
        )

    table_lines.append("  " + "-" * 70)
    grand_quantity_string = (
        f"{int(total_accumulator.total_quantity):,d}"
        if total_accumulator.total_quantity.is_integer()
        else f"{total_accumulator.total_quantity:,.2f}"
    )
    table_lines.append(
        f"  {'TOTAL':<26} | {total_accumulator.transfer_count:<10,d} | {grand_quantity_string:<12} | €{total_accumulator.total_amount_eur:<14,.2f}"
    )

    return "\n".join(table_lines)


def process_vat_report(
    report_path: Path,
    price_catalog: Mapping[str, float],
    output_path: Path,
    export_summary: bool = True,
) -> FileProcessingResult:
    """
    Process a single Amazon VAT CSV report using a pre-loaded price catalog.
    
    Populates Column T (unit price) and Column U (unit price * quantity) for FC_TRANSFER rows.
    Leaves Columns W and AD empty for FC_TRANSFER rows.
    """
    if not report_path.exists():
        raise FileNotFoundError(f"VAT report file not found at: {report_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(report_path, mode="r", encoding=DEFAULT_ENCODING, newline="") as infile:
        reader = csv.reader(infile)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"VAT report file is empty: {report_path}")

        column_indices: Mapping[str, int] = {
            column_name.strip(): index for index, column_name in enumerate(header)
        }

        def resolve_column_index(column_header: ColumnHeader, fallback_index: int) -> int:
            return column_indices.get(column_header.value, fallback_index)

        index_transaction_type = resolve_column_index(ColumnHeader.TRANSACTION_TYPE, 5)
        index_asin = resolve_column_index(ColumnHeader.ASIN, 13)
        index_quantity = resolve_column_index(ColumnHeader.QUANTITY, 16)
        index_cost_price = resolve_column_index(ColumnHeader.COST_PRICE_OF_ITEMS, 19)
        index_item_price_vat_excl = resolve_column_index(ColumnHeader.PRICE_OF_ITEMS_AMT_VAT_EXCL, 20)
        index_total_price_vat_excl = resolve_column_index(ColumnHeader.TOTAL_PRICE_OF_ITEMS_AMT_VAT_EXCL, 22)
        index_total_activity_vat_excl = resolve_column_index(ColumnHeader.TOTAL_ACTIVITY_VALUE_AMT_VAT_EXCL, 29)
        index_departure_country = resolve_column_index(ColumnHeader.DEPARTURE_COUNTRY, 62)
        index_arrival_country = resolve_column_index(ColumnHeader.ARRIVAL_COUNTRY, 65)

        total_rows: int = 0
        fc_transfer_count: int = 0
        fc_transfer_updated: int = 0
        missing_asins: Set[str] = set()
        missing_rows_count: int = 0
        total_value_added: float = 0.0

        route_statistics: MutableMapping[RouteKey, RouteMetric] = defaultdict(RouteMetric)
        output_rows: MutableSequence[MutableSequence[str]] = [header]

        for row_cells in reader:
            total_rows += 1
            if not row_cells:
                continue

            if len(row_cells) < len(header):
                row_cells.extend([""] * (len(header) - len(row_cells)))

            transaction_type = (
                row_cells[index_transaction_type].strip().upper()
                if len(row_cells) > index_transaction_type
                else ""
            )

            if transaction_type == TransactionType.FC_TRANSFER:
                fc_transfer_count += 1
                asin = row_cells[index_asin].strip() if len(row_cells) > index_asin else ""
                quantity_string = (
                    row_cells[index_quantity].strip() if len(row_cells) > index_quantity else "1"
                )

                departure_country = (
                    row_cells[index_departure_country].strip().upper()
                    if len(row_cells) > index_departure_country
                    else ""
                ) or UNKNOWN_COUNTRY_CODE

                arrival_country = (
                    row_cells[index_arrival_country].strip().upper()
                    if len(row_cells) > index_arrival_country
                    else ""
                ) or UNKNOWN_COUNTRY_CODE

                route_key = RouteKey(departure_country, arrival_country)

                try:
                    quantity = float(quantity_string) if quantity_string else 1.0
                except ValueError:
                    quantity = 1.0

                if asin in price_catalog:
                    unit_price = price_catalog[asin]
                    total_price = unit_price * quantity

                    row_cells[index_cost_price] = f"{unit_price:.2f}"
                    row_cells[index_item_price_vat_excl] = f"{total_price:.2f}"

                    if index_total_price_vat_excl < len(row_cells):
                        row_cells[index_total_price_vat_excl] = ""
                    if index_total_activity_vat_excl < len(row_cells):
                        row_cells[index_total_activity_vat_excl] = ""

                    fc_transfer_updated += 1
                    total_value_added += total_price
                    route_statistics[route_key].add_transfer(quantity=quantity, amount_eur=total_price)
                else:
                    missing_asins.add(asin)
                    missing_rows_count += 1
                    route_statistics[route_key].add_transfer(quantity=quantity, amount_eur=0.0)

            output_rows.append(row_cells)

    with open(output_path, mode="w", encoding=DEFAULT_ENCODING, newline="") as outfile:
        writer = csv.writer(outfile, quoting=csv.QUOTE_ALL, lineterminator=CSV_LINE_TERMINATOR)
        writer.writerows(output_rows)

    summary_path = (
        output_path.parent / f"{output_path.stem.replace('_processed', '')}_country_summary.csv"
    )
    if export_summary:
        export_country_summary(route_statistics, summary_path)

    return FileProcessingResult(
        report_path=report_path,
        output_path=output_path,
        summary_path=summary_path if export_summary else None,
        total_rows=total_rows,
        fc_transfer_count=fc_transfer_count,
        fc_transfer_updated=fc_transfer_updated,
        missing_asins=sorted(missing_asins),
        missing_rows_count=missing_rows_count,
        total_value_added=total_value_added,
        route_statistics=dict(route_statistics),
    )


def process_batch(
    input_directory: Path,
    price_catalog_path: Path,
    output_directory: Path | None = None,
) -> BatchProcessingResult:
    """
    Process all VAT reports in a folder in batch mode.
    
    Generates individual filled reports, route summaries, and a consolidated batch summary.
    """
    if not input_directory.is_dir():
        raise NotADirectoryError(f"Input path is not a directory: {input_directory}")

    target_output_directory = output_directory if output_directory else input_directory / "processed"
    target_output_directory.mkdir(parents=True, exist_ok=True)

    report_files: MutableSequence[Path] = [
        file_path
        for file_path in input_directory.glob("*.csv")
        if file_path.is_file()
        and file_path.parent != target_output_directory
        and not file_path.name.startswith(".")
        and not file_path.stem.endswith("_country_summary")
    ]

    if not report_files:
        raise FileNotFoundError(f"No .csv files found in directory: {input_directory}")

    report_files.sort()
    price_catalog = load_price_catalog(price_catalog_path)

    batch_summary_path = target_output_directory / "batch_country_summary.csv"
    batch_result = BatchProcessingResult(
        input_directory=input_directory,
        output_directory=target_output_directory,
        batch_summary_path=batch_summary_path,
        price_catalog_path=price_catalog_path,
        catalog_size=len(price_catalog),
        files_count=len(report_files),
        file_results=[],
    )

    mutable_file_results: MutableSequence[FileProcessingResult] = []
    for report_file in report_files:
        output_file = target_output_directory / report_file.name
        file_result = process_vat_report(report_file, price_catalog, output_file, export_summary=True)
        mutable_file_results.append(file_result)

        batch_result.grand_total_rows += file_result.total_rows
        batch_result.grand_fc_transfers += file_result.fc_transfer_count
        batch_result.grand_fc_updated += file_result.fc_transfer_updated
        batch_result.grand_value_added += file_result.total_value_added
        batch_result.all_missing_asins.update(file_result.missing_asins)

        for route_key, metric in file_result.route_statistics.items():
            batch_result.consolidated_routes[route_key].merge_metrics(metric)

    batch_result.file_results = mutable_file_results
    export_country_summary(batch_result.consolidated_routes, batch_summary_path)
    return batch_result


def print_single_summary(result: FileProcessingResult, catalog_size: int, price_catalog_path: Path) -> None:
    """Display single-file execution summary and country route breakdown."""
    print("\n" + "=" * 74)
    print("  AMAZON VAT REPORT PROCESSING COMPLETE")
    print("=" * 74)
    print(f"  Input Report:       {result.report_path}")
    print(f"  Price Catalog:      {price_catalog_path} ({catalog_size} ASINs loaded)")
    print(f"  Output Report:      {result.output_path}")
    if result.summary_path:
        print(f"  Country Summary:    {result.summary_path}")
    print("-" * 74)
    print(f"  Total Rows:         {result.total_rows:,}")
    print(f"  FC_TRANSFER Rows:   {result.fc_transfer_count:,}")
    print(f"  Successfully Filled:{result.fc_transfer_updated:,}")
    print(f"  Total Cost Filled:  €{result.total_value_added:,.2f}")

    if result.missing_asins:
        print("-" * 74)
        print(
            f"  [WARNING] {len(result.missing_asins)} ASIN(s) not found in price catalog ({result.missing_rows_count} rows):"
        )
        for asin in result.missing_asins:
            print(f"    - {asin}")
        print("  (These rows were left with empty price columns)")
    else:
        print(f"  Missing ASINs:      0 (100% matched)")

    print("-" * 74)
    print("  DEPARTURE COUNTRY x ARRIVAL COUNTRY SUMMARY (FC_TRANSFER):")
    print(format_route_table(result.route_statistics))
    print("=" * 74 + "\n")


def print_batch_summary(result: BatchProcessingResult) -> None:
    """Display batch execution summary with per-file table and consolidated routes."""
    print("\n" + "=" * 74)
    print("  BATCH PROCESSING COMPLETE")
    print("=" * 74)
    print(f"  Input Directory:    {result.input_directory}")
    print(f"  Output Directory:   {result.output_directory}")
    print(f"  Consolidated CSV:   {result.batch_summary_path}")
    print(f"  Price Catalog:      {result.price_catalog_path} ({result.catalog_size} ASINs loaded)")
    print(f"  Files Processed:    {result.files_count}")
    print("-" * 74)
    print(f"  {'Filename':<30} | {'Rows':<7} | {'FC Rows':<8} | {'Filled Value':<14}")
    print("-" * 74)

    for item in result.file_results:
        filename = item.report_path.name
        if len(filename) > 28:
            filename = filename[:25] + "..."
        print(
            f"  {filename:<30} | {item.total_rows:<7,d} | {item.fc_transfer_updated:<8,d} | €{item.total_value_added:<13,.2f}"
        )

    print("-" * 74)
    print(f"  GRAND TOTALS:")
    print(f"    Total Rows:       {result.grand_total_rows:,}")
    print(f"    FC_TRANSFER Rows: {result.grand_fc_transfers:,}")
    print(f"    Filled Rows:      {result.grand_fc_updated:,}")
    print(f"    Total Cost Added: €{result.grand_value_added:,.2f}")

    if result.all_missing_asins:
        print("-" * 74)
        print(f"  [WARNING] {len(result.all_missing_asins)} ASIN(s) not found in price catalog:")
        for asin in sorted(result.all_missing_asins):
            print(f"    - {asin}")
        print("  (Unmatched rows were left with empty price columns)")
    else:
        print(f"    Missing ASINs:    0 (100% matched)")

    print("-" * 74)
    print("  CONSOLIDATED DEPARTURE x ARRIVAL SUMMARY (ALL FILES):")
    print(format_route_table(result.consolidated_routes))
    print("=" * 74 + "\n")


def prompt_for_path(prompt_text: str) -> Path:
    """Prompt user interactively for a file or directory path, sanitizing drag-and-drop input."""
    while True:
        try:
            user_input = input(f"{prompt_text}: ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\nOperation cancelled by user.")
            sys.exit(0)

        if not user_input:
            print("  Please provide a valid path.")
            continue

        sanitized_input = user_input.strip("'\"").replace(r"\ ", " ")
        resolved_path = Path(sanitized_input).expanduser().resolve()

        if not resolved_path.exists():
            print(f"  Error: Path not found at '{resolved_path}'. Please try again.")
            continue
        return resolved_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fill missing prices for FC_TRANSFER rows and calculate Departure x Arrival country sums in Amazon VAT Reports."
    )
    parser.add_argument(
        "--vat-report",
        dest="input_path",
        type=Path,
        help="Path to an Amazon VAT report CSV file or a folder containing reports",
    )
    parser.add_argument(
        "--price-catalog",
        dest="price_catalog_path",
        type=Path,
        help="Path to the Excel price catalog (.xlsx)",
    )
    parser.add_argument(
        "--output",
        dest="output_path",
        type=Path,
        default=None,
        help="Custom destination output path (file for single mode, directory for batch mode)",
    )

    args = parser.parse_args()

    input_path = args.input_path.expanduser().resolve() if args.input_path else None
    price_catalog_path = args.price_catalog_path.expanduser().resolve() if args.price_catalog_path else None
    output_path = args.output_path.expanduser().resolve() if args.output_path else None

    # Interactive prompts if paths are missing
    if not input_path or not price_catalog_path:
        print("\n" + "=" * 74)
        print("   Amazon VAT Report - FC_Transfer Price Automation & Country Summary")
        print("=" * 74)
        print(" Tip: You can drag and drop a file or folder into this window.\n")

        if not input_path:
            input_path = prompt_for_path("1. Enter or Drag & Drop the VAT report CSV file or folder with reports")
        if not price_catalog_path:
            price_catalog_path = prompt_for_path("2. Enter or Drag & Drop the Excel price catalog (.xlsx)")

    try:
        if input_path.is_dir():
            batch_result = process_batch(
                input_directory=input_path,
                price_catalog_path=price_catalog_path,
                output_directory=output_path,
            )
            print_batch_summary(batch_result)
        else:
            price_catalog = load_price_catalog(price_catalog_path)
            default_output_path = input_path.parent / f"{input_path.stem}_processed{input_path.suffix}"
            target_output_path = output_path if output_path else default_output_path
            file_result = process_vat_report(input_path, price_catalog, target_output_path, export_summary=True)
            print_single_summary(file_result, len(price_catalog), price_catalog_path)
    except Exception as error:
        print(f"\n[ERROR] {error}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
