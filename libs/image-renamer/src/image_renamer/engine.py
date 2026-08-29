"""Amazon ASIN Image Duplicator & Renamer Engine.

Duplicates a collection of base template images (<SUFFIX>.<ext>) across a target list of ASINs,
generating an organized folder structure for each ASIN:
    <output_dir>/<ASIN>/<ASIN>.<SUFFIX>.<ext>

Supports loading ASINs from text strings, .txt files, .csv files, and .xlsx Excel catalogs.
Provides both high-performance file copying on disk (CLI) and in-memory WebAssembly packaging (Web).
"""

from __future__ import annotations

import argparse
import csv
import io
import logging
import os
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from collections.abc import Sequence

logger: logging.Logger = logging.getLogger(__name__)

SUPPORTED_IMAGE_EXTENSIONS: frozenset[str] = frozenset({
    "jpg",
    "jpeg",
    "png",
    "webp",
    "tif",
    "tiff",
    "gif",
})

ASIN_REGEX_PATTERN: re.Pattern[str] = re.compile(r"^[A-Z0-9]{10}$")
EXPECTED_SPLIT_PARTS: int = 2


@dataclass(slots=True)
class ImageEntry:
    """Represents a source template image and its extracted suffix/extension."""

    original_name: str
    suffix: str
    extension: str


@dataclass(slots=True)
class TargetImageFile:
    """Represents a generated target file for an ASIN."""

    asin: str
    source_filename: str
    target_folder: str
    target_filename: str
    target_relative_path: str


@dataclass(slots=True)
class DuplicationResult:
    """Summary of the image duplication execution."""

    asins_count: int
    source_images_count: int
    total_created_files: int
    output_directory: Path
    asins: Sequence[str]
    created_files: Sequence[TargetImageFile]


def extract_suffix(filename: str) -> tuple[str, str]:
    """Extract the suffix and extension from an image filename.

    Args:
        filename: Name or path of the image file.

    Returns:
        Tuple containing (suffix_name, extension).

    Examples:
        - "MAIN.jpg" -> ("MAIN", "jpg")
        - "PT01.png" -> ("PT01", "png")
        - "01.jpeg"  -> ("01", "jpeg")
        - "B089WJC6Z4.MAIN.jpg" -> ("MAIN", "jpg")
    """
    raw_path = Path(filename)
    extension = raw_path.suffix.lstrip(".").lower()
    stem = raw_path.stem

    parts = stem.split(".", 1)
    if len(parts) == EXPECTED_SPLIT_PARTS and ASIN_REGEX_PATTERN.match(parts[0].upper()):
        suffix = parts[1].strip()
    else:
        parts_underscore = stem.split("_", 1)
        if len(parts_underscore) == EXPECTED_SPLIT_PARTS and ASIN_REGEX_PATTERN.match(
            parts_underscore[0].upper()
        ):
            suffix = parts_underscore[1].strip()
        else:
            suffix = stem.strip()

    return suffix, extension


def parse_asins_from_text(text: str) -> list[str]:
    """Parse, clean, and deduplicate ASINs from a multi-line or delimited text string.

    Args:
        text: Raw text string containing ASIN tokens.

    Returns:
        Deduplicated list of cleaned, uppercase ASIN strings.
    """
    tokens = re.split(r"[\r\n,;\t\s]+", text)
    seen: set[str] = set()
    asins: list[str] = []

    for token in tokens:
        cleaned = token.strip().upper()
        if not cleaned or cleaned == "ASIN":
            continue
        if cleaned not in seen:
            seen.add(cleaned)
            asins.append(cleaned)

    return asins


def parse_asins_from_csv(csv_content: str | bytes) -> list[str]:
    """Extract ASINs from CSV content, automatically locating the ASIN header column.

    Args:
        csv_content: CSV text or raw UTF-8 encoded bytes.

    Returns:
        List of deduplicated uppercase ASIN strings.
    """
    text = csv_content.decode("utf-8-sig") if isinstance(csv_content, bytes) else csv_content
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []

    header = [col.strip().lower() for col in rows[0]]
    asin_col_idx = 0
    start_row = 0

    if "asin" in header:
        asin_col_idx = header.index("asin")
        start_row = 1
    elif any("asin" in h for h in header):
        asin_col_idx = next(i for i, h in enumerate(header) if "asin" in h)
        start_row = 1

    seen: set[str] = set()
    asins: list[str] = []

    for row in rows[start_row:]:
        if len(row) > asin_col_idx:
            val = row[asin_col_idx].strip().upper()
            if val and val != "ASIN" and val not in seen:
                seen.add(val)
                asins.append(val)

    return asins


def parse_asins_from_excel(file_bytes_or_path: bytes | Path) -> list[str]:
    """Extract ASINs from an Excel .xlsx workbook.

    Args:
        file_bytes_or_path: File bytes or Path to the .xlsx workbook.

    Returns:
        List of deduplicated uppercase ASIN strings.
    """
    if isinstance(file_bytes_or_path, bytes):
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(file_bytes_or_path), data_only=True, read_only=True
        )
    else:
        wb = openpyxl.load_workbook(filename=file_bytes_or_path, data_only=True, read_only=True)

    target_sheet = wb.active
    asin_col_idx = 0
    asins: list[str] = []

    if target_sheet is not None:
        first_row = next(target_sheet.iter_rows(values_only=True), None)
        if first_row:
            for idx, cell in enumerate(first_row):
                if cell and "asin" in str(cell).strip().lower():
                    asin_col_idx = idx
                    break

        seen: set[str] = set()
        for row in target_sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > asin_col_idx:
                cell_val = row[asin_col_idx]
                if cell_val is not None:
                    val = str(cell_val).strip().upper()
                    if val and val != "ASIN" and val not in seen:
                        seen.add(val)
                        asins.append(val)

    wb.close()
    return asins


def parse_asins(source: str | bytes | Path) -> list[str]:
    """Unified ASIN parser supporting strings, .txt paths, .csv, and .xlsx.

    Args:
        source: Text string, byte payload, or filesystem Path.

    Returns:
        List of deduplicated uppercase ASIN strings.
    """
    if isinstance(source, Path):
        ext = source.suffix.lower()
        if ext == ".xlsx":
            return parse_asins_from_excel(source)
        content = source.read_text(encoding="utf-8-sig")
        return parse_asins_from_csv(content) if ext == ".csv" else parse_asins_from_text(content)

    if isinstance(source, bytes):
        if source.startswith(b"PK\x03\x04"):
            return parse_asins_from_excel(source)
        try:
            return parse_asins_from_csv(source)
        except (UnicodeDecodeError, csv.Error):
            return parse_asins_from_text(source.decode("utf-8-sig", errors="ignore"))

    return parse_asins_from_text(str(source))


def generate_image_manifest(
    image_filenames: Sequence[str],
    asins: Sequence[str],
) -> list[TargetImageFile]:
    """Compute the target file list and folder structure for all ASINs.

    Args:
        image_filenames: List of source template image names.
        asins: List of ASIN strings.

    Returns:
        List of TargetImageFile records.
    """
    entries: list[ImageEntry] = []
    for fname in image_filenames:
        base_name = Path(fname).name
        ext = Path(base_name).suffix.lstrip(".").lower()
        if ext in SUPPORTED_IMAGE_EXTENSIONS or not ext:
            suffix, extension = extract_suffix(base_name)
            entries.append(
                ImageEntry(original_name=base_name, suffix=suffix, extension=extension or "jpg")
            )

    manifest: list[TargetImageFile] = []
    for asin in asins:
        for entry in entries:
            target_filename = f"{asin}.{entry.suffix}.{entry.extension}"
            manifest.append(
                TargetImageFile(
                    asin=asin,
                    source_filename=entry.original_name,
                    target_folder=asin,
                    target_filename=target_filename,
                    target_relative_path=f"{asin}/{target_filename}",
                )
            )

    return manifest


def duplicate_images(
    images_dir: Path,
    asins: Sequence[str],
    output_dir: Path,
    *,
    use_hardlinks: bool = False,
) -> DuplicationResult:
    """Duplicate template images across each ASIN into output_dir/<ASIN>/<ASIN>.<SUFFIX>.<ext>.

    Args:
        images_dir: Folder containing template source images.
        asins: Collection of ASIN strings.
        output_dir: Destination folder.
        use_hardlinks: Whether to create OS hardlinks instead of copying data.

    Returns:
        DuplicationResult summarizing created files.

    Raises:
        NotADirectoryError: If images_dir is not a directory.
        ValueError: If asins collection is empty.
        FileNotFoundError: If images_dir contains no supported image files.
    """
    if not images_dir.is_dir():
        msg = f"Images source directory not found: {images_dir}"
        raise NotADirectoryError(msg)

    if not asins:
        msg = "No valid ASINs provided for image duplication."
        raise ValueError(msg)

    output_dir.mkdir(parents=True, exist_ok=True)

    source_files: list[Path] = [
        f
        for f in sorted(images_dir.iterdir())
        if f.is_file()
        and not f.name.startswith(".")
        and f.suffix.lstrip(".").lower() in SUPPORTED_IMAGE_EXTENSIONS
    ]

    if not source_files:
        msg = f"No supported images found in directory: {images_dir}"
        raise FileNotFoundError(msg)

    image_names = [f.name for f in source_files]
    manifest = generate_image_manifest(image_names, asins)

    file_map = {f.name: f for f in source_files}
    created_files: list[TargetImageFile] = []

    for item in manifest:
        asin_folder = output_dir / item.target_folder
        asin_folder.mkdir(parents=True, exist_ok=True)
        dest_path = asin_folder / item.target_filename
        source_path = file_map.get(item.source_filename)

        if source_path:
            if use_hardlinks:
                try:
                    if dest_path.exists():
                        dest_path.unlink()
                    os.link(source_path, dest_path)
                except OSError:
                    shutil.copy2(source_path, dest_path)
            else:
                shutil.copy2(source_path, dest_path)
            created_files.append(item)

    logger.info(
        "Successfully created %d image files across %d ASINs in %s",
        len(created_files),
        len(asins),
        output_dir,
    )

    return DuplicationResult(
        asins_count=len(asins),
        source_images_count=len(source_files),
        total_created_files=len(created_files),
        output_directory=output_dir,
        asins=asins,
        created_files=created_files,
    )


def main() -> None:
    """CLI Entry point for the Amazon ASIN Image Duplicator & Renamer Tool."""
    logging.basicConfig(level=logging.INFO, format="  [%(levelname)s] %(message)s")

    desc = (
        "Duplicate and rename a set of template images across an ASIN list: "
        "<output>/<ASIN>/<ASIN>.<SUFFIX>.<ext>"
    )
    parser = argparse.ArgumentParser(description=desc)
    parser.add_argument(
        "--images",
        dest="images_dir",
        type=Path,
        required=True,
        help="Path to folder containing template images (<SUFFIX>.<ext>)",
    )
    parser.add_argument(
        "--asins",
        dest="asins_source",
        type=str,
        required=True,
        help="List of ASINs, or path to .txt, .csv, or .xlsx file",
    )
    parser.add_argument(
        "--output",
        dest="output_dir",
        type=Path,
        default=Path("output_images"),
        help="Destination directory for ASIN folders (default: ./output_images)",
    )
    parser.add_argument(
        "--hardlinks",
        action="store_true",
        help="Use OS hardlinks instead of copying files (instant, 0 extra disk space)",
    )

    args = parser.parse_args()

    asins_path = Path(args.asins_source).expanduser()
    asins = parse_asins(asins_path) if asins_path.exists() else parse_asins(args.asins_source)

    if not asins:
        print("Error: No ASINs found in input.")
        sys.exit(1)

    print("\n" + "=" * 64)
    print("  AMAZON ASIN IMAGE DUPLICATOR & RENAMER")
    print("=" * 64)
    print(f"  Source Images:   {args.images_dir.resolve()}")
    print(f"  ASINs Count:     {len(asins)}")
    print(f"  Output Folder:   {args.output_dir.resolve()}")
    print(f"  Link Mode:       {'Hardlink (Instant)' if args.hardlinks else 'Copy'}")
    print("-" * 64)

    try:
        result = duplicate_images(
            images_dir=args.images_dir.expanduser().resolve(),
            asins=asins,
            output_dir=args.output_dir.expanduser().resolve(),
            use_hardlinks=args.hardlinks,
        )
        print(
            f"  ✓ Success! Generated {result.total_created_files} images for "
            f"{result.asins_count} ASINs."
        )
        print(f"  Location: {result.output_directory}")
        print("=" * 64 + "\n")
    except (OSError, ValueError, RuntimeError):
        logger.exception("Image duplication failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
